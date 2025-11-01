# requirements_consistency_agent.py
"""
Universal Requirements Consistency Agent
Handles URS/FRS/DS/CS in PDF, DOCX, XLSX, TXT
Outputs a CSV traceability report.
"""

import re
import pandas as pd
from pathlib import Path
from tqdm import tqdm
from sentence_transformers import SentenceTransformer, util

# File parsers
from unstructured.partition.pdf import partition_pdf
from docx import Document
import openpyxl

# ==========================================================
# 🧩 Utility Functions
# ==========================================================

def extract_text_from_file(path):
    ext = Path(path).suffix.lower()
    text_blocks = []

    if ext == ".pdf":
        elements = partition_pdf(filename=path, strategy="hi_res")
        text_blocks = [el.text.strip() for el in elements if hasattr(el, "text") and el.text.strip()]

    elif ext == ".docx":
        doc = Document(path)
        text_blocks = [p.text.strip() for p in doc.paragraphs if p.text.strip()]

    elif ext == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and cell.strip():
                        text_blocks.append(cell.strip())

    elif ext == ".txt":
        text_blocks = Path(path).read_text(encoding="utf-8").splitlines()

    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return [t for t in text_blocks if len(t) > 3]


def segment_requirements(text_blocks):
    """Split paragraphs into requirement-like items"""
    requirements = []
    for block in text_blocks:
        items = re.split(r'(?:\n|^)(?:[0-9A-Za-z]+[\.\-\)]\s+)', block)
        for it in items:
            cleaned = it.strip()
            if len(cleaned) > 10 and "shall" in cleaned.lower() or "should" in cleaned.lower() or "must" in cleaned.lower():
                requirements.append(cleaned)
    return list(set(requirements))


# ==========================================================
# 🔍 Matching Engine
# ==========================================================

def match_docs(doc_a, doc_b, name_a, name_b, threshold=0.65):
    model = SentenceTransformer("all-MiniLM-L6-v2")

    emb_a = model.encode(doc_a, convert_to_tensor=True, normalize_embeddings=True)
    emb_b = model.encode(doc_b, convert_to_tensor=True, normalize_embeddings=True)

    matches = []
    for i, ra in enumerate(tqdm(doc_a, desc=f"Matching {name_a} → {name_b}")):
        sims = util.cos_sim(emb_a[i], emb_b)[0]
        best_idx = int(sims.argmax())
        best_score = float(sims[best_idx])

        status = "OK" if best_score >= threshold else ("Partial" if best_score >= threshold - 0.1 else "Missing")

        matches.append({
            f"{name_a}_REQ": ra[:200],
            f"{name_b}_REQ": doc_b[best_idx][:200],
            "Similarity": round(best_score, 3),
            "Status": status
        })
    return matches


# ==========================================================
# 📄 Main Function
# ==========================================================

def run_traceability(urs_file, frs_file, ds_file, cs_file):
    docs = {
        "URS": urs_file,
        "FRS": frs_file,
        "DS": ds_file,
        "CS": cs_file,
    }

    extracted = {}
    for name, f in docs.items():
        print(f"Extracting from {f} ...")
        text_blocks = extract_text_from_file(f)
        extracted[name] = segment_requirements(text_blocks)
        print(f"  → Found {len(extracted[name])} potential requirements.\n")

    # Pairwise matching
    all_matches = []
    for (a, b) in [("URS", "FRS"), ("FRS", "DS"), ("DS", "CS")]:
        result = match_docs(extracted[a], extracted[b], a, b)
        all_matches.extend(result)

    df = pd.DataFrame(all_matches)
    df.to_csv("traceability_report.csv", index=False)
    print("\n✅ Traceability report generated → traceability_report.csv")


# ==========================================================
# 🚀 Entry Point
# ==========================================================
if __name__ == "__main__":
    run_traceability(
        urs_file="URS.pdf",
        frs_file="FRS.docx",
        ds_file="DS.xlsx",
        cs_file="CS.txt"
    )
